import openpyxl

# 顧客マスタのブック、シート
wb_master = openpyxl.load_workbook("顧客マスタ.xlsx")
ws_master = wb_master["Sheet1"]

# 売上データのブック、シート
wb_data = openpyxl.load_workbook("売上データ_202007.xlsx")
ws_data = wb_data["Sheet1"]

#顧客マスタの全データリスト
customer_list = []

for row in ws_master.iter_rows(min_row=2):
	if row[0].value is None:
		break
	value_list = []
	for c in row:
		value_list.append(c.value)
	customer_list.append(value_list)

# 集計結果を入れるリスト
result_list = []

# 顧客ごとに処理
for customer in customer_list:
	customer_id = customer[0]
	customer_name = customer[1]

	# 売上件数のカウント
	count_sales = 0
	#売上データの検索
	for row in ws_data.iter_rows(min_row=4):
		# 検索条件(売上データの顧客IDが一致)
		if row[1].value == customer_id:
			count_sales = count_sales + 1

	if count_sales > 0:
		# 顧客ID、顧客名称、売上件数の３項目を追加
		result_list.append([customer_id, customer_name, count_sales])

# 集計結果用シートを追加
ws_new = wb_data.create_sheet(title="顧客別売上件数")
# ヘッダー書き込み
ws_new.append(["顧客ID", "顧客名称", "売上件数"])
# 集計結果書き込み
for result in result_list:
	ws_new.append(result)

id_list = []
# 集計結果書き込み
for result in result_list:
	ws_new.append(result)
	# 集計した顧客IDのリストを作成
	id_list.append(result[0])

# 集計されていない顧客を検索
for row in ws_data.iter_rows(min_row=4):
	# 検索条件(集計した顧客IDのリストに含まれない)
	if row[1].value not in id_list:
		# 集計されていない顧客IDと顧客名称を表示
		print(row[1].value, row[2].value)

# 別名で保存
wb_data.save("売上データ_202007_顧客別売上件数.xlsx")
