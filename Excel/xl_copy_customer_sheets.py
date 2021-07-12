import openpyxl

# 顧客マスタのブック、シート
wb_master = openpyxl.load_workbook("顧客マスタ.xlsx")
ws_master = wb_master["Sheet1"]

# 売上データのブック、シート
wb_data = openpyxl.load_workbook("売上データ_202007.xlsx", data_only=True)
ws_data = wb_data["Sheet1"]

# 顧客マスタの全データリスト
customer_list = []

for row in ws_master.iter_rows(min_row=2):
	if row[0].value is None:
		break
	value_list = []
	for c in row:
		value_list.append(c.value)
	customer_list.append(value_list)

# 顧客ごとに処理
for customer in customer_list:
	customer_id = customer[0]
	customer_name = customer[1]
	# 売上データの検索
	data_list = []
	for row in ws_data.iter_rows(min_row=4):
		value_list = []
		for c in row:
			value_list.append(c.value)
		# 検索条件(売上データの顧客IDが一致)
		if value_list[1] == customer_id:
			data_list.append(value_list)

	if len(data_list) > 0:
		# 顧客名称をシート名にしてシート追加
		ws_new = wb_data.create_sheet(title=customer_name)
		# ヘッダー書き込み
		ws_new.append(["売上日", "顧客ID", "顧客名称", "商品名", "単価", "数量", "計"])
		# データ書き込み
		row_num = 2
		for data in data_list:
			ws_new.append(data)
			ws_new.cell(row_num,1).number_format = "yyyy/m/d"
			row_num = row_num + 1

# 別ブックで保存
wb_data.save("売上データ_202007_顧客別シート.xlsx")
