
import openpyxl

wb = openpyxl.load_workbook("売上データ.xlsx")
ws = wb["4月売上"]

c = ws["A1"]
print(c.coordinate)
print(c.row)
print(c.column)
c2 = ws.cell(1, 1)
print(c2.coordinate)
print(c2.row)
print(c2.column)
