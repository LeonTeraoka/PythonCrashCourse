import openpyxl

wb = openpyxl.load_workbook("売上データ.xlsx")
# 先頭に挿入
ws_new1 = wb.create_sheet(index=0)
# 末尾に挿入
ws_new2 = wb.create_sheet()

# シートの名前を表示
print(ws_new1.title)
print(ws_new2.title)

wb.save("売上データ_シート挿入.xlsx")
