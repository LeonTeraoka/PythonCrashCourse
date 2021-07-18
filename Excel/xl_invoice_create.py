import openpyxl
import datetime

# 顧客マスタのブック、シート
wb_master = openpyxl.load_workbook("顧客マスタ.xlsx")
ws_master = wb_master["Sheet1"]

#売上データのブック、シート
wb_data = openpyxl.load_workbook("売上データ_202007.xlsx", data_only=True)
ws_data = wb_data["Sheet1"]

# 請求書ひな型のブック、シート
wb_inv = openpyxl.load_workbook("請求書ひな型.xlsx")
ws_inv = wb_inv["Sheet1"]


# 顧客マスタの全データリスト
customer_list = []

for row in ws_master.iter_rows(min_row=2):
	if row[0].value is None:
		break
	value_list = []
	for c in row:
		value_list.append(c.value)
	customer_list.append(value_list)

# 顧客ごとに処理
for customer in customer_list:
	customer_id = customer[0]
	customer_name = customer[1]
	# 売上データの検索
	data_list = []
	for row in ws_data.iter_rows(min_row=4):
		value_list = []
		for c in row:
			value_list.append(c.value)
		# 検索条件(売上データの顧客IDが一致)
		if value_list[1] == customer_id:
			data_list.append(value_list)

	if len(data_list) > 0:
		# 確認
		print(data_list[0])
		print(data_list[-1])
		#ここから請求書作成の処理を行う(後半のコード)
		