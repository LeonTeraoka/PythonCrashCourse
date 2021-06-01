import openpyxl

wb = openpyxl.load_workbook("売上データ_6月入力.xlsx")

# 1つ目のシートかどうか
is_first_sheet = True

# このリストに全シートから読み取ったデータをまとめる
row_list = []

# ブックの中のすべてのシートを処理
for ws in wb.worksheets:

	if is_first_sheet:
		start_row = 1
	else:
		start_row = 4

	for row in ws.iter_rows(min_row=start_row):
		# ヘッダーより下で空行になったら読み込み修了
		if row[0].row > 3 and row[0].value is None:
			break
		value_list = []
		for c in row:
			value_list.append(c.value)
		row_list.append(value_list)

	is_first_sheet = False

# データを転記する新しいシート
ws_new = wb.create_sheet(title="第1四半期売上")

# 書き込み時の行番号
row_num = 1

# 新しいシートに1行ずつデータを書き込む
for row in row_list:
	# 1行分のデータを書き込む
	ws_new.append(row)

	# データ部分のA列に日付の表示形式を設定
	if row_num > 3:
		ws_new.cell(row_num, 1).number_format = "yyyy/m/d"

	row_num = row_num + 1

# 別名でブック保存
wb.save("売上データ_第1四半期売上.xlsx")
