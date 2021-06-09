import csv
import openpyxl
from pathlib import Path
import datetime

# 新規ブック作成
wb = openpyxl.Workbook()

# 「売上月別」フォルダー内CSVファイルの読み込み
header_num = 3
for file in Path("売上月別").glob("*.csv"):
	#新規シートを追加
	ws = wb.create_sheet(file.stem)
	# CSVファイルを開く
	f = open(file)
	reader = csv.reader(f)
	# CSVファイルを1行ずつ読み込む
	for row in reader:
		if reader.line_num <= header_num:
			# ヘッダー行はそのまま追加
			ws.append([row[0], row[1], row[5]])
			continue

		# データ行は適宜データ型を変換
		dt = datetime.datetime.strptime(row[0], "%Y/%m/%d")
		customer = row[1]
		sales = int(row[5])
		# シートにデータ行を追加
		ws.append([dt, customer, sales])
		# 日付のセルは書式を指定する
		ws.cell(reader.line_num, 1).number_format = "yyyy/m/d"
	f.close()

# ブック作成時の既存シートを取り除く
ws_first = wb.worksheet[0]
wb.remove(ws_first)

# ブック保存
wb.save("第1四半期売上.xlsx")