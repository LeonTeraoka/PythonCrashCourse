import openpyxl

# 顧客マスタのブック、シート
wb_master = openpyxl.load_workbook("顧客マスタ.xlsx")
ws_master = wb_master["Sheet1"]

# 売上データのブック、シート
wb_data = openpyxl.load_workbook("売上データ_202007.xlsx", data_only=True)
ws_data = wb_data["Sheet1"]

#顧客マスタの全データリスト
customer_list = []

for row in ws_master.iter_rows(min_row=2):
	if row[0].value is None:
		break
	value_list = []
	for c in row:
		value_list.append(c.value)
	customer_list.append(value_list)

# 集計結果を入れるリスト
result_list = []

# 顧客ごとに処理
for customer in customer_list:
	customer_id = customer[0]
	customer_name = customer[1]
	# 売上の集計
	sum_sales = 0
	#売上データの検索
	for row in ws_data.iter_rows(min_row=4):
		# 検索条件(売上データの顧客IDが一致)
		if row[1].value == customer_id:
			# 売上の額は[G列(インデックス =6)]
			sum_sales = sum_sales + row[6].value

	if sum_sales >0:
		# 顧客ID、顧客名称、売上件数の３項目を追加
		result_list.append([customer_id, customer_name, sum_sales])

# 集計結果用シートを追加
ws_new = wb_data.create_sheet(title="顧客別売上")
# ヘッダー書き込み
ws_new.append(["顧客ID", "顧客名称", "売上計"])
# 集計結果書き込み
for result in result_list:
	ws_new.append(result)

# 別名で保存
wb_data.save("売上データ_202007_顧客別売上合計.xlsx")
