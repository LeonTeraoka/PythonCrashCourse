import openpyxl

# 顧客マスタのブック、シート
wb_master = openpyxl.load_workbook("顧客マスタ.xlsx")
ws_master = wb_master["Sheet1"]

# 売上データのブック、シート
wb_data = openpyxl.load_workbook("売上データ_202007.xlsx", data_only=True)
ws_data = wb_data["Sheet1"]

#顧客マスタの全データリスト
customer_list = []

for row in ws_master.iter_rows(min_row=2):
	if row[0].value is None:
		break
	value_list = []
	for c in row:
		value_list.append(c.value)
	customer_list.append(value_list)

# データ行番号
row_num = 3
#列名追加
ws_data["H" + str(row_num)].value = "当社営業担当"

#売上データを1行ずつ処理
for row in ws_data.iter_rows(min_row=4):
	row_num = row_num + 1

	#顧客 ID: 売上データの[B列](インデックス=1)
	customer_id = row[1].value

	#顧客の検索
	for customer in customer_list:
		# 検索条件(顧客IDの一致)
		if customer[0] == customer_id:
			# 当社営業担当を[H列]に追加
			ws_data["H" + str(row_num)].value = customer[5]
			break

# 別名で保存
wb_data.save("売上データ_202007_営業担当あり.xlsx")
