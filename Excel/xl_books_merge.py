import openpyxl
from pathlib import Path

# 各支店ブックの読み込み
wb_list = []
for file in Path("売上全支店").glob("*.xlsx"):
	wb = openpyxl.load_workbook(file)
	wb_list.append(wb)

# 保存先の新しいブック
wb_new = openpyxl.Workbook()

# 1つ目のブックからシート名のリストを取得
sheet_names = wb_list[0].sheetnames

# シート名ごとに処理
for sheet_name in sheet_names:
	# このシートに書き込む
	ws_new = wb_new.create_sheet(sheet_name)

	# 1つ目のブックかどうか
	is_first_book = True

	#このリストに全ブックから読み取ったデータをまとめる
	row_list = []

	# 各支店のブックごとにデータ読み込み
	for wb in wb_list:
		# 読み込むシート
		ws = wb[sheet_name]

		if is_first_book:
			start_row = 1
		else:
			start_row = 4

		#シートから1行ずつデータを読み込む
		for row in ws.iter_rows(min_row=start_row):
			# ヘッダーより下で空行になったら読み込み修了
			if row[0].row > 3 and row[0].value is None:
				break
			value_list = []
			for c in row:
				value_list.append(c.value)
			row_list.append(value_list)

		is_first_book = False

	# 書き込み時の行番号
	row_num = 1

	# シートに1行ずつデータを書き込む
	for row in row_list:
		# 1行ぶんのデータを書き込む
		ws_new.append(row)

		#データの部分のA列に日付の表示形式を設定
		if row_num > 3:
			ws_new.cell(row_num, 1).number_format = "yyyy/m/d"

		row_num = row_num + 1

# 作成時の既存シートを取り除く
ws_first = wb_new.worksheets[0]
wb_new.remove(ws_first)

wb_new.save("売上データ_全国.xlsx")
